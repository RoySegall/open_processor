import os
import json
import openpyxl


class ExcelParser():
    def __init__(self, logger):
        self._logger = logger
        self._is_israel = None

    def parse_file(self, file_path):
        """

        :param file_path:
        :return:
        """
        # Load in the workbook file
        workbook = openpyxl.load_workbook(file_path)

        if not workbook:
            self._logger.error("Failed to load excel file")
            return False

        for sheet_name in workbook.sheetnames:
            if workbook[sheet_name].title == 'סכום נכסי הקרן':
                # :todo: need parse this sheet ?
                continue
            self.parse_sheet(workbook[sheet_name])

    def parse_sheet(self, sheet, start_row=0, start_column=2):
        """

        :param sheet:
        :param start_row:
        :param start_column:
        :return:
        """
        current_row = start_row
        current_column = start_column
        current_cell = None
        row_data = None

        data = {
            "metadata": {
                "name   ": sheet.title

            },
            "data": {}
        }

        while current_cell not in ['שם נ"ע', 'שם המנפיק/שם נייר ערך']:
            if current_cell:
                metadata = self._get_metadata(data=row_data)
                if metadata:
                    data["metadata"][metadata[0]] = metadata[1]

            current_row += 1
            row_data = ExcelAdapter.get_entire_row(sheet=sheet,
                                                   row=current_row,
                                                   min_column=current_column)

            if row_data:
                # strip all spaces from start and end string
                current_cell = row_data[0].strip()
            else:
                current_cell = None
        else:
            # get fields names
            fields_name = ExcelAdapter.get_entire_row(sheet=sheet,
                                                      row=current_row,
                                                      min_column=start_column)

            fields_len = len(fields_name)

        empty_len = 0
        current_cell = ""
        data_row = []
        while current_cell != '* בעל ענין/צד קשור':

            if empty_len > 5:
                self._logger.warn("max empty row")
                break

            # get next row
            current_row += 1
            data_row = ExcelAdapter.get_entire_row(sheet=sheet,
                                                   row=current_row,
                                                   min_column=start_column,
                                                   max_column=fields_len+start_column)

            # Check if is empty row or first cell is empty
            if not data_row or not data_row[0]:
                empty_len += 1
                continue

            current_cell = data_row[0]

            if current_cell.find('סה"כ') != -1:
                self._parse_total_field(current_cell)
            else:
                data["data"][data_row[0]] = {
                    "total": self._total_data,
                    "israel": self._is_israel
                }
                for i in range(1, fields_len):
                    try:
                        data["data"][data_row[0]][fields_name[i]] = data_row[i]
                    except IndexError as ex:
                        self._logger.error("Failed {0} {1}".format(ex, fields_name))


        print(data)

        file = "/tmp/{0}".format(data["metadata"]["name"])
        if file:
            self._save_to_json_file(file_path=file, data=data)

    def _get_metadata(self, data):
        """

        :param data: list of data
        :return:
        """
        first_cell = data[0]
        if not first_cell:
            self._logger.error("No data in first cell")
            return None, None

        finder = first_cell.find(":")
        # find func return -1 when not find
        if finder != -1:
            # Check if the colon char is not last data char (the mean the data in the first cell)
            if len(first_cell) > finder:
                return first_cell[:finder], first_cell[finder+1:]

        # check if len of data is bigger than one
        elif len(data) > 1:
            return first_cell, data[1]

    def _parse_total_field(self, data):
        self._total_data = data.strip('סה"כ')

        find = lambda search: False if self._total_data.find(search) == -1 else True
        string_found = lambda words_list: True if len(list(filter(find, words_list))) else False

        if string_found(words_list=['ישראל', 'בארץ']):
            self._is_israel = True
            print("{0} is israel".format(data))
        elif string_found(words_list=['מט"ח', 'חוץ לארץ', 'חו"ל']):
            self._is_israel = False
            print("{0} is not israel".format(data))

        print(self._total_data)

    def _save_to_json_file(self, file_path, data):
        if not os.path.isdir(file_path):
            self._logger.error("folder not exists {0}".format(file_path))
        try:
            with open(file_path, "w") as outfile:
                json.dump(data, outfile)
            return True
        except Exception as ex:
            self._logger.error("Failed to write json file {0}".format(ex))
            return False


class ExcelAdapter:
    @staticmethod
    def get_cell(sheet, row, column):
        """
        Get value for row
        :param sheet:
        :param row:
        :param column:
        :return:
        """
        try:
            return sheet.cell(row=row, column=column).value
        except Exception as ex:
            raise Exception("Failed to read cell {0}".format(ex))

    @staticmethod
    def get_entire_row(sheet, row, min_column=1, max_column=None):
        """

        :param sheet:
        :param row:
        :param min_column:
        :return:
        """
        cell_data = None
        row_data = []
        column = min_column

        data_exists = lambda: True if cell_data else False
        is_not_max_column = lambda: not(max_column == column)

        #
        if max_column:
            check = is_not_max_column
        else:
            check = data_exists

        cell_data = ExcelAdapter.get_cell(sheet=sheet, column=column, row=row)

        while check():
            row_data.append(cell_data)
            column += 1
            cell_data = ExcelAdapter.get_cell(sheet=sheet, column=column, row=row)

        return row_data


class FakeLogger:
    def error(self, msg):
        print("error {0}".format(msg))

    def info(self, msg):
        print("info {0}".format(msg))

    def warn(self,msg):
        print("warring {0}".format(msg))

if __name__ == '__main__':
    logger = FakeLogger()
    process_xl = ExcelParser(logger=logger)
    process_xl.parse_file(file_path="test32.xlsx")
